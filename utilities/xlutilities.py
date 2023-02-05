import openpyxl

def getrowcount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    row=sheet.max_row
    return row

def getcolumncount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    column=sheet.max_column
    return column

def readdata(file,sheetname,rownum,colnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    data=sheet.cell(row=rownum,column=colnum).value
    return data

def writedata(file,sheetname,rownum,colnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(row=rownum,column=colnum).value=data
    workbook.save(file)